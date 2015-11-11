import pandas as pd
import openpyxl as pyxl
import os

def grab_values(excel_sheet):
    wb = pyxl.load_workbook(excel_sheet, data_only=True)
    ws_list = wb.get_sheet_names()
    named_ranges = [named_range.name for named_range in wb.get_named_ranges()]
    variable_dict = {}
    for named_range in wb.get_named_ranges():
        try:
            sheet_name = named_range.destinations[0][0].title
            cell_name = named_range.destinations[0][1].replace('$','')
            variable_dict[named_range.name] = {}
            variable_dict[named_range.name]['value'] = wb[sheet_name][cell_name].value
            variable_dict[named_range.name]['sheet'] = sheet_name
            variable_dict[named_range.name]['cell'] = cell_name
        except:
            pass
    return variable_dict
    
def run_program():
    list_of_dicts = {}
    location_dicts = {}
    for sheet in os.listdir(os.getcwd()):
        try:
            # location_dicts[sheet] = grab_values("{}\\{}".format(os.getcwd(),sheet)) # windows
            # temp_dict = grab_values("{}\\{}".format(os.getcwd(),sheet)) # windows
            # Uncomment above by highlighting the line and ctrl or cmd / removing the leftmost # if on windows and comment mac below
            location_dicts[sheet] = grab_values(sheet) # mac
            temp_dict = grab_values(sheet) # mac
            df_dict = {}
            for variable, dictionary in temp_dict.items():
                try:
                    df_dict[variable] = dictionary['value']
                except:
                    pass
            list_of_dicts[sheet] = df_dict
        except:
            pass
    df = pd.DataFrame(list_of_dicts).T
    df.to_csv('named_range_database.csv', index_label='workbook')
    return df, location_dicts

def edit_all_files(new_csv, location_dict):
    csv_dict = pd.read_csv(new_csv, fill_na="").T.to_dict()
    for workbook_dict in csv_dict.values():
        wb_filename = workbook_dict['workbook']
        # wb = pyxl.load_workbook("{}\\{}".format(os.getcwd(), wb_filename)) # windows
        wb = pyxl.load_workbook(wb_filename, data_only=True) # mac
        for variable, value in workbook_dict.items(): 
            if variable != "workbook":
                try:
                    sheet_name = location_dict[wb_filename][variable]['sheet']
                    cell_name = location_dict[wb_filename][variable]['cell']
                    wb[sheet_name][cell_name] = value
                except:
                    pass
        # wb.save("{}\\{}".format(os.getcwd(), wb_filename)) # windows       
        wb.save("new_{}".format(wb_filename)) # mac
        print('Saved {}'.format(wb_filename))

df, excel_dicts = run_program()
df

edit_all_files('named_range_database.csv', excel_dicts) # run this if you want to edit original excel files
